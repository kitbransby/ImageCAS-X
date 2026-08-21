import json
import os
import numpy as np
import SimpleITK as sitk


def load_volume(path: str) -> tuple[np.ndarray, sitk.Image]:
    """(array [x,y,z], sitk image for the geometry)."""
    img = sitk.ReadImage(path)
    arr = sitk.GetArrayFromImage(img).transpose(2, 1, 0)  # [z,y,x] -> [x,y,z]
    return arr, img


def load_mask(path: str) -> tuple[np.ndarray, sitk.Image]:
    """(mask [x,y,z], sitk image)."""
    img = sitk.ReadImage(path)
    arr = sitk.GetArrayFromImage(img).transpose(2, 1, 0)
    return arr.astype(np.uint8), img


# Coronary lumen label scheme: 0 = background, 1-14 = branch lumen.
LUMEN_BACKGROUND_LABEL = 0


def binarise_lumen(mask: np.ndarray,
                   background_label: int = LUMEN_BACKGROUND_LABEL) -> np.ndarray:
    """Multi-label mask -> binary lumen: every label but background."""
    return (mask != background_label).astype(np.uint8)


def save_mask(arr: np.ndarray, reference_img: sitk.Image, out_path: str):
    """Save a mask, copying geometry from reference."""
    out = sitk.GetImageFromArray(arr.transpose(2, 1, 0).astype(np.uint8))
    out.CopyInformation(reference_img)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    sitk.WriteImage(out, out_path)


def save_prob(arr: np.ndarray, reference_img: sitk.Image, out_path: str):
    """Save a probability map, copying geometry from reference (--save-probs)."""
    out = sitk.GetImageFromArray(arr.transpose(2, 1, 0).astype(np.float32))
    out.CopyInformation(reference_img)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    sitk.WriteImage(out, out_path)


def save_results(results: dict, out_path: str):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2)


def load_descriptors(path: str, id_column: str = "Scan ID") -> dict:
    """Scan-level descriptor table (.xlsx or .csv) as {scan_id: {column: value}}.

    IDs and values are stringified and stripped so a numeric "1" in the spreadsheet
    matches the "1" in a filelist; blank cells become "" and read as missing.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError(
                f"Reading {path} needs openpyxl (`pip install openpyxl`), or convert "
                f"the descriptor table to .csv and point at that instead."
            ) from e
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]
        wb.close()
    elif ext == ".csv":
        import csv
        with open(path, newline="") as f:
            rows = [r for r in csv.reader(f)]
    else:
        raise ValueError(f"Unsupported descriptor file type '{ext}': {path} (expected .xlsx or .csv)")

    rows = [r for r in rows if r and any(c is not None and str(c).strip() for c in r)]
    if not rows:
        raise ValueError(f"Descriptor file is empty: {path}")

    header = ["" if c is None else str(c).strip() for c in rows[0]]
    if id_column not in header:
        raise ValueError(
            f"Descriptor file {path} has no '{id_column}' column. Columns found: {header}"
        )
    id_idx = header.index(id_column)

    def _cell(v):
        if v is None:
            return ""
        # Excel stores integer cells as floats; "1.0" must key as "1".
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    table = {}
    for row in rows[1:]:
        row = list(row) + [None] * (len(header) - len(row))
        scan_id = _cell(row[id_idx])
        if not scan_id:
            continue
        table[scan_id] = {
            col: _cell(val)
            for col, val in zip(header, row)
            if col and col != id_column
        }
    return table


def resolve_scan_path(base_dir: str, scan_id: str, extension: str) -> str:
    path = os.path.join(base_dir, f"{scan_id}{extension}")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Scan file not found: {path}")
    return path


def load_vtk_centerline(path: str) -> tuple[np.ndarray, dict]:
    """VTK polydata centerline as (points in LPS mm, {array name: values})."""
    try:
        import vtk
        from vtk.util.numpy_support import vtk_to_numpy
    except ImportError:
        raise ImportError("vtk is required to load centerline VTK files: pip install vtk")

    reader = vtk.vtkPolyDataReader()
    reader.SetFileName(path)
    # Without these the reader returns only the file's *active* scalars, so a
    # centerline carrying several point arrays (segment_label, end_points, ...) would
    # silently come back with just one of them.
    reader.ReadAllScalarsOn()
    reader.ReadAllFieldsOn()
    reader.Update()
    poly = reader.GetOutput()

    if poly.GetNumberOfPoints() == 0:
        return np.zeros((0, 3), dtype=np.float64), {}

    points = vtk_to_numpy(poly.GetPoints().GetData()).astype(np.float64)

    scalars = {}
    pd = poly.GetPointData()
    for i in range(pd.GetNumberOfArrays()):
        arr = pd.GetArray(i)
        if arr is not None:
            scalars[arr.GetName()] = vtk_to_numpy(arr)

    return points, scalars


def load_vtk_centerline_branches(path: str) -> list:
    """Per-branch point arrays in walk order, from the file's line/cell connectivity.

    `load_vtk_centerline` reads only GetPoints() and returns one flat array whose
    order need not match a walk along the vessel. That is fine for nearest-point
    lookups but would let anything interpolating between consecutive points draw
    spurious shortcuts between points that are not actually adjacent.

    Falls back to a single branch in raw point order, with a warning, if the file has
    no line cells at all.
    """
    try:
        import vtk
        from vtk.util.numpy_support import vtk_to_numpy
    except ImportError:
        raise ImportError("vtk is required to load centerline VTK files: pip install vtk")

    reader = vtk.vtkPolyDataReader()
    reader.SetFileName(path)
    reader.Update()
    poly = reader.GetOutput()

    if poly.GetNumberOfPoints() == 0:
        return []

    all_points = vtk_to_numpy(poly.GetPoints().GetData()).astype(np.float64)

    lines = poly.GetLines()
    if lines is None or lines.GetNumberOfCells() == 0:
        print(f"    [warn] {path}: no line/cell connectivity found — falling back to "
              f"raw point order as a single branch (walk order not verified)")
        return [all_points]

    branches = []
    lines.InitTraversal()
    id_list = vtk.vtkIdList()
    while lines.GetNextCell(id_list):
        ids = [id_list.GetId(i) for i in range(id_list.GetNumberOfIds())]
        if len(ids) > 0:
            branches.append(all_points[ids])
    return branches


def read_image_geometry(path: str):
    """A NIfTI file's header without its pixel data. The reader exposes the same
    getters as a real sitk.Image, so it is a drop-in argument for the functions
    below."""
    reader = sitk.ImageFileReader()
    reader.SetFileName(path)
    reader.ReadImageInformation()
    return reader


def resampled_geometry(sitk_img_or_reader, target_spacing: float) -> sitk.Image:
    """The isotropic target grid, rounded exactly as Resample computes it for the
    offline caches — so this reproduces that grid from a header alone, without paying
    for an interpolation pass. Returns an empty image; use only its header."""
    src = sitk_img_or_reader
    orig_spacing = src.GetSpacing()
    orig_size = src.GetSize()
    new_size = [int(round(orig_size[i] * orig_spacing[i] / target_spacing)) for i in range(3)]
    geom = sitk.Image(new_size, src.GetPixelID())
    geom.SetSpacing([target_spacing] * 3)
    geom.SetOrigin(src.GetOrigin())
    geom.SetDirection(src.GetDirection())
    return geom


def resample_image_to_spacing(sitk_img: sitk.Image, target_spacing: float,
                              interpolator=sitk.sitkNearestNeighbor,
                              default_value: float = 0.0) -> sitk.Image:
    """Resample onto the grid `resampled_geometry` describes, so voxel-count-sized
    patches span consistent physical sizes across scans whatever a scan's original
    spacing. Defaults to nearest-neighbour, for label masks."""
    orig_spacing = sitk_img.GetSpacing()
    orig_size = sitk_img.GetSize()
    new_size = [int(round(orig_size[i] * orig_spacing[i] / target_spacing)) for i in range(3)]
    resampler = sitk.ResampleImageFilter()
    resampler.SetOutputSpacing([target_spacing] * 3)
    resampler.SetSize(new_size)
    resampler.SetOutputDirection(sitk_img.GetDirection())
    resampler.SetOutputOrigin(sitk_img.GetOrigin())
    resampler.SetInterpolator(interpolator)
    resampler.SetDefaultPixelValue(default_value)
    return resampler.Execute(sitk_img)


def centerline_points_to_mask(branches: list, ref_img: sitk.Image) -> np.ndarray:
    """Rasterise centerline branches onto ref_img's grid as a thin 1-voxel line.

    Interior voxels between each branch's own consecutive points are interpolated, so
    no gaps appear where points are spaced more than a voxel apart — but interpolation
    never crosses a branch boundary, so two unrelated branches are never bridged.
    """
    size = ref_img.GetSize()  # (X, Y, Z)
    mask = np.zeros(size, dtype=np.uint8)
    if not branches:
        return mask

    origin = np.array(ref_img.GetOrigin())
    spacing = np.array(ref_img.GetSpacing())
    direction = np.array(ref_img.GetDirection()).reshape(3, 3)
    inv_direction = np.linalg.inv(direction)

    def _to_voxel(pts_mm: np.ndarray) -> np.ndarray:
        rel = pts_mm - origin[None, :]
        return (inv_direction @ rel.T).T / spacing[None, :]

    all_vox = []
    for branch_mm in branches:
        if len(branch_mm) == 0:
            continue
        vox = _to_voxel(np.asarray(branch_mm, dtype=np.float64))
        if len(vox) > 1:
            segments = [vox]
            for p0, p1 in zip(vox[:-1], vox[1:]):
                n_steps = max(int(np.ceil(np.linalg.norm(p1 - p0))), 1)
                t = np.linspace(0.0, 1.0, n_steps + 1)[1:-1]  # interior points only
                if len(t) > 0:
                    segments.append(p0[None, :] + t[:, None] * (p1 - p0)[None, :])
            vox = np.concatenate(segments, axis=0)
        all_vox.append(vox)

    if not all_vox:
        return mask
    vox = np.concatenate(all_vox, axis=0)

    idx = np.round(vox).astype(np.int64)
    shape = np.array(size)
    in_bounds = np.all((idx >= 0) & (idx < shape[None, :]), axis=1)
    idx = idx[in_bounds]
    mask[idx[:, 0], idx[:, 1], idx[:, 2]] = 1
    return mask
